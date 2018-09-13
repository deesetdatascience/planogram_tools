import re
import cv2 as cv
import numpy as np
import openpyxl
from pathlib import Path

gc_code_re = re.compile(
    r"([A-Z]{5}$|[0-9]{12}$|[0-9]{7}$|[0-9]{6}$|[0-9]+CMBLACKPEGS$|HOTSPOT$)"
)


def read_excel_planogram(plano_path, product_images_dir=None, cell_w=10, cell_h=10):
    """Reads in an excel planogram file as a python object"""

    # Conver paths to Path objects
    plano_path = Path(plano_path)

    # Load the workbook
    wb = openpyxl.load_workbook(plano_path)
    ws = wb.active

    # Get cell width and height if in excel sheet and not specified
    if cell_w == cell_h == 10:
        cell_w = cell_h = ws.cell(1, 1).value

    # Find the maximum extent of the planogram
    excel_width, excel_height = ws.max_column, ws.max_row

    # Set the background colour
    bg_colour = "#{}".format(ws.cell(1, 1).fill.fgColor.rgb[2:])

    # Define the base info for the planogram
    plan = {
        "store": "wilko",
        "bay": plano_path.with_suffix("").name,
        "bg_colour": bg_colour,
        "slots": [],
        "planogram_path": plano_path,
    }

    # Define the width and height of the planogram image
    plan["width"], plan["height"] = (
        (excel_width + 1) * cell_w,
        (excel_height + 1) * cell_h,
    )

    # Add each slot to the planogram object

    for i, merged_cell in enumerate(
        sorted(ws.merged_cells.ranges, key=lambda x: (x.bounds[1], x.bounds[0]))
    ):

        slot = {"index": i}

        # Define the bounds of the current slot
        bounds = merged_cell.bounds
        slot["left"] = (bounds[0]) * cell_w
        slot["top"] = (bounds[1]) * cell_h
        slot["width"] = (bounds[2] - bounds[0] + 1) * cell_w
        slot["height"] = (bounds[3] - bounds[1] + 1) * cell_h

        # Get the Code that corresponds to the current slot
        contents = ws.cell(bounds[1], bounds[0]).value
        if contents:

            # Remove any slashes, as we will use this for folder names
            contents = contents.replace("/", "")

            # Try and extract just the NAV code from the last brackets
            match = re.match(r"^.*\((.*)\)$", contents)
            if match:
                contents = match.group(1)

            slot["name"] = contents

        # Find an image for the current slot if necessary
        if product_images_dir:
            image_paths = list((Path(product_images_dir)).glob(f"*{contents}/*"))
            if image_paths:
                image_path = image_paths[0]
                slot["image_path"] = image_path

        # Add the current slot definition to the planogram object
        plan["slots"].append(slot)

    return plan
